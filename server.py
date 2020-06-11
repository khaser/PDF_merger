#!/usr/bin/python3.7

import logging
import random
import zipfile
import os
import openpyxl
import sys
from flask import Flask, render_template, request, flash, redirect, send_from_directory
from PyPDF2 import PdfFileReader, PdfFileWriter


ROOT = os.path.split(os.path.abspath(__file__))[0]
DOWNLOAD_FOLDER = os.path.join(ROOT, 'tmp')
UPLOAD_FOLDER = os.path.join(ROOT, 'result')

application = Flask(__name__)
application.secret_key = "abracadabra"
log = logging.getLogger(__name__)
logging.basicConfig(level = logging.DEBUG, format = "> %(asctime)-15s %(levelname)-8s || %(message)s")

def allowedFile(filename, allowedExtensions):
	return '.' in filename and filename.rsplit('.', 1)[1] in allowedExtensions

def downloadFolder(files, sess, saveFullPath = 0):
	path = os.path.join(DOWNLOAD_FOLDER, sess)
	os.makedirs(path, exist_ok = True)
	cnt = 0
	for data in files:
		if (allowedFile(data.filename, 'pdf')):
			cnt += 1
			if (saveFullPath):
				os.makedirs(os.path.join(path, os.path.split(data.filename)[0]), exist_ok = True)
				data.save(os.path.join(path, data.filename))
			else:
				data.save(os.path.join(path, os.path.split(data.filename)[1]))
	return cnt

@application.route("/")
def index():
	return render_template("index.html")

@application.route("/favicon.ico")
def favicon():
	return send_from_directory(os.path.join(application.root_path, 'static'), 'favicon.ico',\
		mimetype='image/vnd.microsoft.icon')

@application.route('/uploadFolder', methods=['POST', 'GET'])
def uploadFolder():
	sess = "id" + str(random.randint(0, 10**30))
	cntFiles = downloadFolder(request.files.getlist("file[]"), sess)
	mergeDir(os.path.join(DOWNLOAD_FOLDER, sess), os.path.join(UPLOAD_FOLDER, sess))
	rmDir(os.path.join(DOWNLOAD_FOLDER, sess))
	return send_from_directory(UPLOAD_FOLDER, sess + '.pdf')

@application.route('/upload', methods=['POST', 'GET'])
def upload():
	files = request.files.getlist("files[]")
	sess = "id" + str(random.randint(0, 10**30))
	path = os.path.join(DOWNLOAD_FOLDER, sess)
	os.makedirs(path, exist_ok=True)
	cnt = 0
	for data in files:
		cnt += 1
		s = str(cnt).zfill(7)
		data.save(os.path.join(path, s + ".pdf"))
	merge(sess)
	rmDir(path)
	return send_from_directory(UPLOAD_FOLDER, sess + '.pdf')

@application.route('/uploadRecursive', methods=['POST', 'GET'])
def uploadRecursive():
	sess = "id" + str(random.randint(0, 10**30))
	cntFiles = downloadFolder(request.files.getlist("file[]"), sess, saveFullPath = 1)
	mergeRecursive(sess)
	rmDir(os.path.join(DOWNLOAD_FOLDER, sess))
	zipFile = os.path.join(UPLOAD_FOLDER, sess + '.zip')
	zipDir(os.path.join(UPLOAD_FOLDER, sess), zipFile)
	return send_from_directory(UPLOAD_FOLDER, os.path.split(zipFile)[1]) 

@application.route('/uploadTypes', methods=['POST', 'GET'])
def uploadTypes():
	sess = "id" + str(random.randint(0, 10**30))
	cntFiles = downloadFolder(request.files.getlist("file[]"), sess)
	for file in os.listdir(os.path.join(DOWNLOAD_FOLDER, sess)):
		typeCheck(os.path.join(DOWNLOAD_FOLDER, sess, file), os.path.join(UPLOAD_FOLDER, sess + '.txt'))
	rmDir(os.path.join(DOWNLOAD_FOLDER, sess))
	return send_from_directory(UPLOAD_FOLDER, sess + '.txt', as_attachment=True)

@application.route('/uploadExcel', methods=['POST', 'GET'])
def uploadExcel():
	sess = "id" + str(random.randint(0, 10**30))
	cntFiles = downloadFolder(request.files.getlist("file[]"), sess)
	for data in request.files.getlist("file[]"):
		if (allowedFile(data.filename, 'xlsx')):
			data.save(os.path.join(DOWNLOAD_FOLDER, sess, 'merge.xlsx'))
	mergeExcel(sess)
	rmDir(os.path.join(DOWNLOAD_FOLDER, sess))
	zipFile = os.path.join(UPLOAD_FOLDER, sess + '.zip')
	zipDir(os.path.join(UPLOAD_FOLDER, sess), zipFile)
	return send_from_directory(UPLOAD_FOLDER, os.path.split(zipFile)[1])


def rmDir(dir):
	log.info("Removed directory %s", dir)
	for root, dirs, files in os.walk(dir, topdown=False):
		for name in files:
			os.remove(os.path.join(root, name))
		for name in dirs:
			os.rmdir(os.path.join(root, name))
	os.rmdir(dir)


def zipDir(dir, outpath):
	log.info("Directory %s zipped to %s", dir, outpath)
	zipFile = zipfile.ZipFile(outpath, 'w', zipfile.ZIP_DEFLATED)
	os.chdir(dir)
	for root, dirs, files in os.walk('.'):
		for file in files:
			zipFile.write(os.path.join(root, file))
	zipFile.close()
	os.chdir(ROOT)


def merge(sess):
	output = PdfFileWriter()
	path = os.path.join(DOWNLOAD_FOLDER, sess)
	files = sorted(os.listdir(path))
	for name in files:
		pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
		for i in range(pdfcur.getNumPages()):
			output.addPage(pdfcur.getPage(i))
	outputStream = open(os.path.join(UPLOAD_FOLDER, sess) + '.pdf', "wb")
	output.write(outputStream)	
	outputStream.close()

def mergeDir(path, outpath):
	output = PdfFileWriter()
	files = os.listdir(path)
	for name in files:
		if 'R-LR' in name:
			pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
			for i in range(pdfcur.getNumPages()):
				output.addPage(pdfcur.getPage(i))
	for name in files:
		if 'R-LA' in name:
			pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
			for i in range(pdfcur.getNumPages()):
				output.addPage(pdfcur.getPage(i))
	for name in files:
		if 'R-LB' in name:
			pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
			for i in range(pdfcur.getNumPages()):
				output.addPage(pdfcur.getPage(i))
	outputStream = open(outpath + '.pdf', "wb")
	output.write(outputStream)	
	outputStream.close()

def mergeRecursive(sess):
	os.chdir(os.path.join(DOWNLOAD_FOLDER, sess))
	for i in os.walk('.'):
		path = os.path.normpath(i[0])
		folders = i[1]
		files = [file for file in i[2] if file.endswith('.pdf')]
		if (len(folders) > 0):
			new_folder = os.path.join(UPLOAD_FOLDER, sess, path)
			os.makedirs(new_folder, exist_ok=True)
		if (len(files) > 0):
			mergeDir(path, os.path.join(UPLOAD_FOLDER, sess, path));
	os.chdir(ROOT)


def typeCheck(name, outfile, sheet = ''):
	d1 = {'A0' : [2384, 3370,200], 'A1' : [1684, 2384,150], 'A2' : [1191,1684,120], 'A3' : [842,1191,95], 'A4' : [595,842,70], 'A5' : [420, 595,60], 'A6' : [298, 420,50], 'A7' : [210, 298,40], 'A8' : [147, 210,28],'A9' : [105,147,27], 'A10' : [74 ,105,23]}
	d = {'A0' : [2384, 3370,200], 'A1' : [1684, 2384,150], 'A2' : [1191,1684,120], 'A3' : [842,1191,95], 'A4' : [595,842,70], 'A5' : [420, 595,60], 'A6' : [298, 420,50], 'A7' : [210, 298,40], 'A8' : [147, 210,28],'A9' : [105,147,27], 'A10' : [74 ,105,23]}

	def getType(point):
		x = point[0]
		y = point[1]
		if x > y:
			x, y = y, x
		for i in d:
			if (x >= d[i][0] - d[i][2] and x <= d[i][0] + d[i][2]) or (y >= d[i][1] - d[i][2] and y <= d[i][1] + d[i][2]):
				return i
		return 'unknown'

	for i in d1:
		if 'x' in i:
			continue
		for cnt in range(2,6):
			d[i + 'x' + str(cnt)] = [d[i][0] * cnt, d[i][1] * cnt, d[i][2]]

	fout = open(outfile, 'a')
	with fout as sys.stdout:
		pdfcur = PdfFileReader(open(name, "rb"))
		print('Документ', os.path.split(name)[1], "из листа", sheet, ':')
		d2 = {'A0' : [], 'A1' : [], 'A2' : [], 'A3' : [], 'A4' : [], 'A5' : [], 'A6' : [], 'A7' : [], 'A8' : [],'A9' : [], 'A10' : [], 'unknown' : []}
		for i in d1:
			if 'x' in i:
				continue
			for cnt in range(2,6):
				d2[i + 'x' + str(cnt)] = []
		for i in range(pdfcur.getNumPages()):
			page = pdfcur.getPage(i)
			point = page.mediaBox.upperRight
			d2[getType(point)].append(i + 1)
		noComma = 'kek'
		for i in d2:
			if len(d2[i]) > 0:
				noComma = i
		for i in d2:
			if len(d2[i]) == 0:
				continue
			last = d2[i][0]
			for j in range(1,len(d2[i])):
				if d2[i][j] != d2[i][j-1] + 1:
					if last == d2[i][j-1]:
						print(last, end = ', ', sep = '')
					else:
						print(last,'-',d2[i][j-1], end = ', ', sep = '')
					last = d2[i][j]
			if i == noComma:
				if last == d2[i][len(d2[i]) - 1]:
					print(last, '-', i, sep = '')
				else:
					print(last, '-', d2[i][len(d2[i]) - 1], '-', i, sep = '')

			else:
				if last == d2[i][len(d2[i]) - 1]:
					print(last	, '-', i, end = ',', sep = '')
					print()
				else:
					print(last, '-', d2[i][len(d2[i]) - 1], '-', i, end = ',', sep = '')
					print()

def mergeExcel(sess):
	path = os.path.join(DOWNLOAD_FOLDER, sess)
	outpath = os.path.join(UPLOAD_FOLDER, sess)

	def addPDF(name, stream):
		pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
		for i in range(pdfcur.getNumPages()):
			stream.addPage(pdfcur.getPage(i))

	wb = openpyxl.load_workbook(filename = os.path.join(path, 'merge.xlsx'))
	sheet = wb[wb.sheetnames[0]]

	not_found = []

	os.makedirs(outpath)
	for sheet_name in wb.sheetnames:
		sheet = wb[sheet_name]
		empty_cnt = 0
		d = {}
		cur = "trash"

		for i in range(1, 100500):
			if (empty_cnt > 5):
				break;
			file = sheet.cell(row=i, column=1)
			if (file.value == None):
				empty_cnt += 1;
				continue;
			empty_cnt = 0;
			if (file.font.b == True):
				cur = file.value
				continue;
			file.value = str(file.value) + '.pdf'
			if (os.path.exists(os.path.join(path, file.value))):
				d.setdefault(cur, []).append(file.value) 
			else:
				not_found.append(file.value)

		for i in d.items():
			fout, files = i[0], i[1]
			fout = os.path.join(outpath, sheet_name, str(fout))
			
			output = PdfFileWriter()
			for item in files:
				addPDF(item, output)
		
			os.makedirs(os.path.split(fout)[0], exist_ok=True)
			outputStream = open(fout + '.pdf', "wb")
			output.write(outputStream)
			outputStream.close()
		report = open(os.path.join(outpath, "Отчет.txt"), 'a')
		for i in d.items():
			print('Файл', i[0], 'был склеен из:', file = report)
			print(*i[1], sep = '\n', end = '\n\n', file = report)
		report.close()


	report = open(os.path.join(outpath, "Не найденные файлы.txt"), 'w') 
	if (len(not_found) != 0):
		print("Мы не нашли след. файлы, но все равно склеили без них:", file = report)
	else:
		print("Все файлы на месте", file = report)
	print(*not_found, sep = '\n', file = report)
	report.close()

	for sheet in wb.sheetnames:
		for file in sorted(os.listdir(os.path.join(outpath, sheet))):
			typeCheck(os.path.join(outpath, sheet, file), os.path.join(outpath, "Форматы файлов.txt"), sheet = sheet)




def main():
	application.run(host='0.0.0.0', debug=True, port=80)

if __name__ == "__main__":
	main()
