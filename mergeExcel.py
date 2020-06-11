#!/usr/bin/python3.7

from PyPDF2 import PdfFileReader, PdfFileWriter
import os
import sys
import openpyxl

global path
path = os.path.abspath(__file__)
path = os.path.split(path)[0]
sess = sys.argv[1]
outpath = os.path.join(path, 'result', sess)
path = os.path.join(path, 'tmp', sess)

def addPDF(name, stream):
    pdfcur = PdfFileReader(open(os.path.join(path, name), "rb"))
    for i in range(pdfcur.getNumPages()):
        stream.addPage(pdfcur.getPage(i))

wb = openpyxl.load_workbook(filename = os.path.join(path, 'merge.xlsx'))
sheet = wb[wb.sheetnames[0]]

global not_found
not_found = []

os.makedirs(outpath)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    empty_cnt = 0
    d = {}
    cur = "trash"

    for i in range(1, 100500):
        if (empty_cnt > 5):
            break;
        file = sheet.cell(row=i, column=1)
        if (file.value == None):
            empty_cnt += 1;
            continue;
        empty_cnt = 0;
        if (file.font.b == True):
            cur = file.value
            continue;
        file.value = str(file.value) + '.pdf'
        if (os.path.exists(os.path.join(path, file.value))):
            d.setdefault(cur, []).append(file.value) 
        else:
            not_found.append(file.value)

    for i in d.items():
        fout, files = i[0], i[1]
        fout = os.path.join(outpath, sheet_name, str(fout))
        
        output = PdfFileWriter()
        for item in files:
            addPDF(item, output)
    
        os.makedirs(os.path.split(fout)[0], exist_ok=True)
        outputStream = open(fout + '.pdf', "wb")
        output.write(outputStream)
        outputStream.close()
    report = open(os.path.join(outpath, "Отчет.txt"), 'a')
    for i in d.items():
        print('Файл', i[0], 'был склеен из:', file = report)
        print(*i[1], sep = '\n', end = '\n\n', file = report)
    report.close()




report = open(os.path.join(outpath, "Не найденные файлы.txt"), 'w') 
if (len(not_found) != 0):
    print("Мы не нашли след. файлы, но все равно склеили без них:", file = report)
else:
    print("Все файлы на месте", file = report)
print(*not_found, sep = '\n', file = report)
report.close()

d1 = {'A0' : [2384, 3370,200], 'A1' : [1684, 2384,150], 'A2' : [1191,1684,120], 'A3' : [842,1191,95], 'A4' : [595,842,70], 'A5' : [420, 595,60], 'A6' : [298, 420,50], 'A7' : [210, 298,40], 'A8' : [147, 210,28],'A9' : [105,147,27], 'A10' : [74 ,105,23]}
d = {'A0' : [2384, 3370,200], 'A1' : [1684, 2384,150], 'A2' : [1191,1684,120], 'A3' : [842,1191,95], 'A4' : [595,842,70], 'A5' : [420, 595,60], 'A6' : [298, 420,50], 'A7' : [210, 298,40], 'A8' : [147, 210,28],'A9' : [105,147,27], 'A10' : [74 ,105,23]}
def getType(point):
    x = point[0]
    y = point[1]
    if x > y:
        x, y = y, x
    for i in d:
        if (x >= d[i][0] - d[i][2] and x <= d[i][0] + d[i][2]) or (y >= d[i][1] - d[i][2] and y <= d[i][1] + d[i][2]):
            return i
    return 'unknown'

for i in d1:
    if 'x' in i:
        continue
    for cnt in range(2,6):
        d[i + 'x' + str(cnt)] = [d[i][0] * cnt, d[i][1] * cnt, d[i][2]]

report = open(os.path.join(outpath, 'Типы файлов.txt'), 'w')
sys.stdout = report 
for folder in os.walk(outpath):
    for name in folder[2]:
        if (name.rsplit('.', 1)[1] != 'pdf'):
            continue
        pdfcur = PdfFileReader(open(os.path.join(outpath, folder[0], name), "rb"))
        print('Папка', os.path.split(folder[0])[1], 'Документ', name, ':')
        d2 = {'A0' : [], 'A1' : [], 'A2' : [], 'A3' : [], 'A4' : [], 'A5' : [], 'A6' : [], 'A7' : [], 'A8' : [],'A9' : [], 'A10' : [], 'unknown' : []}
        for i in d1:
            if 'x' in i:
                continue
            for cnt in range(2,6):
                d2[i + 'x' + str(cnt)] = []
        for i in range(pdfcur.getNumPages()):
            page = pdfcur.getPage(i)
            point = page.mediaBox.upperRight
            d2[getType(point)].append(i + 1)
        noComma = 'kek'
        for i in d2:
            if len(d2[i]) > 0:
                noComma = i
        for i in d2:
            if len(d2[i]) == 0:
                continue
            last = d2[i][0]
            for j in range(1,len(d2[i])):
                if d2[i][j] != d2[i][j-1] + 1:
                    if last == d2[i][j-1]:
                        print(last, end = ', ', sep = '')
                    else:
                        print(last,'-',d2[i][j-1], end = ', ', sep = '')
                    last = d2[i][j]
            if i == noComma:
                if last == d2[i][len(d2[i]) - 1]:
                    print(last, '-', i, sep = '')
                else:
                    print(last, '-', d2[i][len(d2[i]) - 1], '-', i, sep = '')

            else:
                if last == d2[i][len(d2[i]) - 1]:
                    print(last, '-', i, end = ',', sep = '')
                    print()
                else:
                    print(last, '-', d2[i][len(d2[i]) - 1], '-', i, end = ',', sep = '')
                    print()
report.close()
